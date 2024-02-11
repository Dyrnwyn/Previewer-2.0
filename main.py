import json
import os.path
import platform
from re import compile as re_compile
from static_method import get_file_list_in_path_used_re_template, show_message_box_with_question
from sys import argv
from PyQt5 import QtWidgets
from PyQt5.QtCore import Qt
from os import path, sep, mkdir, scandir, getcwd
from previewer2_main_window import Ui_MainWindow
from previewer2_logick import Previewer
from openpyxl import load_workbook
from shutil import copy


class MainApp(QtWidgets.QMainWindow, Ui_MainWindow):
    font_path = ""
    if platform.system() == "Linux":
        font_path = '{0}{1}Font{2}'.format(getcwd(), sep, sep)
    elif platform.system() == "Windows":
        font_path = '{0}{1}{2}{3}'.format(os.path.expandvars(r'%LOCALAPPDATA%'), sep +'Microsoft',
                                                                                       sep + 'Windows',
                                                                                       sep + 'Fonts' + sep)
    settings_directory = path.expanduser("~") + sep + "Previewer2.0" + sep
    settings = {
        'species': 0,
        'format': 1,
        'template': 2,
        'photo': 3,
        'number': 4,
        'class': 5,
        'last_name': 9,
        'id': 11,
        'summ': 12,
        'holst': False,
        'with_price': False,
        'path': '',
        'object_name': '',
        'font_regular': font_path + 'afuturica.ttf',
        'font_bold': font_path + 'afuturicaextrabold.ttf',
        'font_italic': font_path + 'afuturicaitalic.ttf',
        'psd_files': [],
        'holst_files': [],
        'holst_files_subdir': 'renamed_files',
        'text_for_qr': """ST00012|Name=ООО «Объемный мир»|PersonalAcc=40702810631000007404
                          |BankName=ПАОСБЕРБАНК|BIC=040407627|CorrespAcc=30101810800000000627
                          |PayeeINN=2464105021|KPP=246001001|PersAcc={id}|Sum={summ}""",
        'qr_error_correct': {'index': 3, 'text': '30%'},
        'bottom_text': "vk.com/omfoto_ru\nОбъемныйМир.рф/parents"
    }

    def __init__(self):
        super().__init__()
        self.setupUi(self)

        self.error_plainTextEdit.setVisible(False)
        self.error_plainTextEdit.setEnabled(True)

        self.progressBar.setValue(0)
        self.progressBar.setVisible(False)
        self.label_what_in_work.setVisible(False)

        self.checkBox_holst.clicked.connect(self.holst_check_choice)
        self.checkBox_with_price.clicked.connect(self.price_check_choice)

        self.toolButton_choice_path.clicked.connect(self.set_path_and_object_name)
        self.pushButton_compose.clicked.connect(self.compose_preview)

        self.pushButton_remove_file_settings.clicked.connect(self.set_default_settings)

        self.pushButton_save_settings.clicked.connect(self.save_settings)
        self.toolButton_RegularFont.clicked.connect(self.tool_button_regular_select_file)
        self.toolButton_BoldFont.clicked.connect(self.tool_button_bold_select_file)
        self.toolButton_ItalicFont.clicked.connect(self.tool_button_italic_select_file)

        self.previewer_thread = Previewer()
        self.previewer_thread.started.connect(self.on_start_previewer)
        self.previewer_thread.finished.connect(self.on_finish_previewer)
        self.previewer_thread.what_in_work.connect(self.on_change_what_in_work, Qt.QueuedConnection)
        self.previewer_thread.progress_bar_percent.connect(self.on_change_set_value_pg, Qt.QueuedConnection)
        self.previewer_thread.progress_bar_maximum.connect(self.on_change_maximum_pg, Qt.QueuedConnection)
        self.previewer_thread.missed_files.connect(self.on_change_missed_files)

        if not path.exists(self.settings_directory + sep + 'settings.conf'):
            self.set_default_settings()
        else:
            self.load_settings()

    def set_default_settings(self):
        self.set_default_settings_on_form()
        if not path.exists(self.settings_directory):
            mkdir(self.settings_directory)
        self.save_settings()

    @staticmethod
    def show_message_box(title_text, message_text):
        msg_box = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Warning,
                                        title_text, message_text,
                                        buttons=QtWidgets.QMessageBox.Ok)

        msg_box.exec_()

    def check_font(self):
        fonts = [self.settings['font_regular'],
                 self.settings['font_bold'],
                 self.settings['font_italic']]
        for font in fonts:
            if not path.exists(font):
                self.show_message_box("Ошибка открытия Шрифта", "Не удалось обнаружить шрифт по указанному пути: "
                                      + font)
                return False
        return True

    def on_change_what_in_work(self, msg):
        self.label_what_in_work.setText(msg)

    def on_change_set_value_pg(self, v):
        self.progressBar.setValue(v)

    def on_change_maximum_pg(self, v):
        self.progressBar.setValue(0)
        self.progressBar.setMaximum(v)

    def on_change_missed_files(self, msg):
        if not self.error_plainTextEdit.isVisible():
            self.error_plainTextEdit.setVisible(True)
        self.error_plainTextEdit.insertPlainText(msg)

    def on_start_previewer(self):
        self.error_plainTextEdit.setPlainText("")
        self.error_plainTextEdit.setVisible(False)
        self.progressBar.setVisible(True)
        self.label_what_in_work.setVisible(True)
        self.lineEdit_path.setEnabled(False)
        self.lineEdit_object_name.setEnabled(False)
        self.pushButton_compose.setEnabled(False)
        self.checkBox_holst.setEnabled(False)
        self.checkBox_with_price.setEnabled(False)

    def on_finish_previewer(self):
        self.show_message_box("Завершение операции", "Формирование превью успешно завершено")
        self.lineEdit_path.setEnabled(True)
        self.lineEdit_object_name.setEnabled(True)
        self.pushButton_compose.setEnabled(True)
        self.checkBox_holst.setEnabled(True)
        self.checkBox_with_price.setEnabled(True)

    def compose_preview(self):
        if self.check_data_to_compose_preview():
            if self.settings['holst']:
                self.prepare_files_for_holst_preview()
            else:
                self.find_files_for_preview()
            self.previewer_thread.set_var(self.settings)
            self.previewer_thread.start()

    def check_data_to_compose_preview(self):
        if not self.lineEdit_path.text():
            self.show_message_box("Ошибка", "Не выбрана папка с объектом")
            return False
        elif not self.lineEdit_object_name.text():
            self.show_message_box("Ошибка", "Необходимо указать имя объект")
            return False
        elif not self.check_font():
            return False
        else:
            return True

    def parse_exel_file(self, order_file):
        data_from_sheet_foto = []
        order_file = load_workbook(path.join(self.settings['path'], order_file), data_only=True)
        sheet_foto = order_file['Фото']
        for row in sheet_foto.iter_rows(min_row=11, max_col=12, max_row=2000):
            row_dict = {}
            for cell in row:
                dict_key = ''
                if cell.column_letter == 'A':
                    dict_key = "second_name"
                elif cell.column_letter == 'B':
                    dict_key = "class"
                elif cell.column_letter == 'C':
                    dict_key = "template"
                elif cell.column_letter == 'D':
                    dict_key = "photo"
                elif cell.column_letter == 'E':
                    dict_key = "format"
                elif cell.column_letter == 'F':
                    dict_key = "species"
                elif cell.column_letter == 'G':
                    dict_key = "comment"
                elif cell.column_letter == 'H':
                    dict_key = "number"
                elif cell.column_letter == 'I':
                    dict_key = "price"
                elif cell.column_letter == 'K':
                    dict_key = "order_summ"
                elif cell.column_letter == 'L':
                    dict_key = "actually_get"
                elif cell.column_letter == 'M':
                    dict_key = "summ"
                if cell.value:
                    row_dict[dict_key] = cell.value
                else:
                    row_dict[dict_key] = ""
            data_from_sheet_foto.append(row_dict)
        order_file.close()
        return data_from_sheet_foto

    def create_holst_subdir(self):
        holst_subdir = path.join(self.settings['path'], self.settings['holst_files_subdir'])
        if not path.exists(holst_subdir):
            mkdir(holst_subdir)

    def rename_files_from_order_file(self, data_from_exel):
        self.create_holst_subdir()
        holst_files = []
        file_list = get_file_list_in_path_used_re_template(self.settings["path"], r'.*\.jp.*')
        for row in data_from_exel:
            if row['photo'] == '':
                continue
            i = 0
            compare_pattern = re_compile(row["photo"] + r'\.jp.*')
            row_foto_finded = False
            while i < len(file_list):
                if not row_foto_finded and compare_pattern.fullmatch(file_list[i]):
                    new_file_name = 'п_' + row['format'] + '_0000_' + str(row['photo']) + '_' \
                                    + str(row['number']) + '_' + str(row['class']) + '_V_id_' \
                                    + str(row['order_summ']) + '_' + str(row['second_name']) + '_V.jpg'
                    copy(path.join(self.settings["path"], file_list[i]),
                         path.join(self.settings["path"], self.settings["holst_files_subdir"], new_file_name))
                    holst_files.append(new_file_name)
                    row_foto_finded = True
                i += 1

        self.settings['holst_files'] = holst_files

    def prepare_files_for_holst_preview(self):
        xls_files = get_file_list_in_path_used_re_template(self.settings['path'], r'.*\.xls.*')
        if len(xls_files) > 1:
            self.show_message_box("Ошибка", "Было найдено два exel файла, необходимо оставить в папке только"
                                            " актуальный файл.")
        elif len(xls_files) == 1:
            data_from_exel = self.parse_exel_file(xls_files[0])
            self.rename_files_from_order_file(data_from_exel)
        else:
            self.show_message_box("Ошибка", "В указанной директории не было найдено, ниодного exel файла.")

    def find_files_for_preview(self):
        refuse_files = []
        file_name_length = 7
        if self.settings['with_price']:
            file_name_length = 12
        file_list = get_file_list_in_path_used_re_template(self.settings['path'], r'.*\.psd')
        psd_files = []
        for file_name in file_list:
            if (len(file_name.split('_'))) < file_name_length:
                refuse_files.append(file_name)
            else:
                psd_files.append(file_name)
        self.settings['psd_files'] = psd_files
        if len(refuse_files) > 0:
            message_text = "В файлах:\n{0}указана не вся информация для формирования превью.\n" \
                           + "Файлы будут пропущены"
            files = ""
            for f in refuse_files:
                files += f + "\n"
            self.show_message_box("Ошибка", message_text.format(files))
        return True

    def set_default_settings_on_form(self):
        self.set_settings_on_form()

    def get_settings_from_form(self):
        self.settings['species'] = self.spinBox_species.value()
        self.settings['format'] = self.spinBox_format.value()
        self.settings['template'] = self.spinBox_template.value()
        self.settings['photo'] = self.spinBox_photo.value()
        self.settings['number'] = self.spinBox_number.value()
        self.settings['class'] = self.spinBox_class.value()
        self.settings['id'] = self.spinBox_id.value()
        self.settings['summ'] = self.spinBox_summ.value()
        self.settings['text_for_qr'] = self.textEdit_qr_text.toPlainText()
        self.settings['qr_error_correct']['index'] = self.comboBox_error_correction_level.currentIndex()
        self.settings['qr_error_correct']['text'] = self.comboBox_error_correction_level.currentText()
        self.settings['font_regular'] = self.lineEdit_RegularFont.text()
        self.settings['font_bold'] = self.lineEdit_BoldFont.text()
        self.settings['font_italic'] = self.lineEdit_ItalicFont.text()
        self.settings['bottom_text'] = self.textEdit_bottom_text.toPlainText()

    def set_settings_on_form(self):
        self.spinBox_summ.setValue(self.settings['species'])
        self.spinBox_format.setValue(self.settings['format'])
        self.spinBox_template.setValue(self.settings['template'])
        self.spinBox_photo.setValue(self.settings['photo'])
        self.spinBox_number.setValue(self.settings['number'])
        self.spinBox_class.setValue(self.settings['class'])
        self.spinBox_id.setValue(self.settings['id'])
        self.spinBox_summ.setValue(self.settings['summ'])
        self.textEdit_qr_text.setText(self.settings['text_for_qr'])
        self.comboBox_error_correction_level.setCurrentIndex(self.settings['qr_error_correct']['index'])
        self.lineEdit_RegularFont.setText(self.settings['font_regular'])
        self.lineEdit_BoldFont.setText(self.settings['font_bold'])
        self.lineEdit_ItalicFont.setText(self.settings['font_italic'])
        self.textEdit_bottom_text.setText(self.settings['bottom_text'])

    def save_settings(self):
        self.get_settings_from_form()
        json_txt = json.dumps(self.settings)
        settings_file = open(self.settings_directory + 'settings.conf', 'w')
        settings_file.write(json_txt)
        settings_file.close()

    def tool_button_regular_select_file(self):
        self.select_file(self.lineEdit_RegularFont, "Выберите файл шрифта 'Regular'")

    def tool_button_bold_select_file(self):
        self.select_file(self.lineEdit_BoldFont, "Выберите файл шрифта 'Bold'")

    def tool_button_italic_select_file(self):
        self.select_file(self.lineEdit_ItalicFont, "Выберите файл шрифта 'Italic'")

    def select_file(self, line_edit, caption_text="Выберите файл шрифта"):
        selected_file = QtWidgets.QFileDialog.getOpenFileName(self.window(), caption=caption_text,
                                                              filter='Fonts(ttf) (*.ttf)',
                                                              initialFilter='Fonts(ttf) (*.ttf)')
        if not selected_file[0] == '':
            line_edit.setText(selected_file[0])

    def load_settings(self):
        if not path.isfile(self.settings_directory + 'settings.conf'):
            self.save_settings()
        settings_file = open(self.settings_directory + 'settings.conf', 'r')
        json_text = settings_file.read()
        settings_file.close()
        error_in_settings = False
        if json_text != '':
            settings = json.loads(json_text)
            for k, v in settings.items():
                self.settings[k] = settings[k]
            if error_in_settings:
                self.show_message_box("Ошибка файла настроек",
                                      "В файле настроек обнаружена ошибка, необходимо проверить настройки")
        self.set_settings_on_form()

    def set_path_and_object_name(self):
        path_text = QtWidgets.QFileDialog.getExistingDirectory()
        self.lineEdit_path.setText(path_text)
        self.settings['path'] = path_text
        self.lineEdit_object_name.setText(path_text.split("/")[-1])
        self.settings['object_name'] = self.lineEdit_object_name.text()

    def holst_check_choice(self):
        if self.checkBox_holst.isChecked():
            self.checkBox_with_price.setCheckState(False)
            self.settings['holst'] = True
        else:
            self.settings['holst'] = False

    def price_check_choice(self):
        if self.checkBox_with_price.isChecked():
            self.checkBox_holst.setCheckState(False)
            self.settings['with_price'] = True
        else:
            self.settings['with_price'] = False


def main():
    app = QtWidgets.QApplication(argv)
    main_window = MainApp()
    main_window.show()
    app.exec_()


if __name__ == '__main__':
    main()
