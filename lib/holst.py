from lib.static_method import show_message_box, create_dir, get_list_of_xls_files, get_list_of_jpeg_files
from openpyxl import load_workbook
from re import compile as re_compile
from shutil import copy
from os import path


def get_prepared_files_for_holst_preview(object_path, subdir_name):
    xls_files = get_list_of_xls_files(object_path)
    if len(xls_files) > 1:
        show_message_box("Ошибка",
                         "Было найдено два exel файла, "
                         "необходимо оставить в папке только актуальный файл.")
    elif len(xls_files) == 1:
        data_from_exel = parse_exel_file(xls_files[0], object_path)
        return rename_files_from_order_file(data_from_exel, subdir_name, object_path)
    else:
        show_message_box("Ошибка",
                         "В указанной директории не было найдено, ниодного exel файла.")


def parse_exel_file(order_file_name, object_path):
    data_from_sheet_foto = []
    order_file = load_workbook(path.join(object_path, order_file_name), data_only=True)
    sheet_foto = order_file['Фото']
    for row in sheet_foto.iter_rows(min_row=11, max_col=12, max_row=2000):
        row_dict = {}
        for cell in row:
            dict_key = ''
            if cell.column_letter == 'A':
                dict_key = "second_name"
            elif cell.column_letter == 'B':
                dict_key = "class"
            elif cell.column_letter == 'C':
                dict_key = "template"
            elif cell.column_letter == 'D':
                dict_key = "photo"
            elif cell.column_letter == 'E':
                dict_key = "format"
            elif cell.column_letter == 'F':
                dict_key = "species"
            elif cell.column_letter == 'G':
                dict_key = "comment"
            elif cell.column_letter == 'H':
                dict_key = "number"
            elif cell.column_letter == 'I':
                dict_key = "price"
            elif cell.column_letter == 'K':
                dict_key = "order_summ"
            elif cell.column_letter == 'L':
                dict_key = "actually_get"
            elif cell.column_letter == 'M':
                dict_key = "summ"
            if cell.value:
                row_dict[dict_key] = cell.value
            else:
                row_dict[dict_key] = ""
        data_from_sheet_foto.append(row_dict)
    order_file.close()
    return data_from_sheet_foto


def rename_files_from_order_file(data_from_exel, holst_subdir, object_path):
    dir_renamed_files = path.join(object_path, holst_subdir)
    create_dir(dir_renamed_files)
    holst_files = []
    file_list = get_list_of_jpeg_files(object_path)
    for row in data_from_exel:
        if row['photo'] == '':
            continue
        i = 0
        compare_pattern = re_compile(row["photo"] + r'\.jp.*')
        row_foto_finded = False
        while i < len(file_list):
            if not row_foto_finded and compare_pattern.fullmatch(file_list[i]):
                new_file_name = 'п_' + row['format'] + '_0000_' + str(row['photo']) + '_' \
                                + str(row['number']) + '_' + str(row['class']) + '_V_id_' \
                                + str(row['order_summ']) + '_' + str(row['second_name']) + '_V.jpg'
                copy(path.join(object_path, file_list[i]), path.join(dir_renamed_files, new_file_name))
                holst_files.append(new_file_name)
                row_foto_finded = True
            i += 1
    return holst_files
